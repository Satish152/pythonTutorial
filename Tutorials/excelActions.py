import openpyxl

filename="../File/excelTest.xlsx"
wb=openpyxl.load_workbook(filename)
ws=wb["TestData"]
print(wb.active.title)
ws["A1"]="Hello"
ws2=wb.create_sheet("scenario2",1)
Test=[["Username","City"],["Satish","Hyderabad"],["Ramesh","Vijayawada"]]

for list in Test:
    ws2.append(list)


ws3=wb.create_sheet("scenario 3",2)
#here inworkbook index starts with 1 instead of 0
for row in range(1,3):  # iteration runs 1,2 and same for second for loop also
    for cell in range(1,3):
        if row==1:
            ws3.cell(row,cell,value="header %s" %(cell))
        else:
            ws3.cell(row,cell,value="Value {value}".format(value=cell))
wb.save(filename)
wb.close()
